import datetime
import re
import sqlite3
import time

from openpyxl import load_workbook


def clear():
    c = conn.cursor()
    c.execute('DELETE from COMPANY')
    conn.commit()


def extract_date(str):
    try:
        matches = re.findall("\d\d", str)
        return time.strftime("%Y-", time.localtime()) + matches[0] + '-' + matches[1] + ' '
    except IndexError:
        matches = re.findall("(\d)月(\d\d)日", str)
        return time.strftime("%Y-", time.localtime()) + '0'+matches[0][0] + '-' + matches[0][1] + ' '


def execute(start_time, end_time, meet_name, people, room):
    print(start_time, end_time, meet_name, people, room)
    c = conn.cursor()
    c.execute('''
    INSERT INTO
        COMPANY
        (START_TIME,END_TIME,MEET_NAME,PEOPLE,ROOM)
    VALUES
       ( ?,?,?,?,?)
    ''', (start_time, end_time, meet_name, people, room))
    conn.commit()


conn = sqlite3.connect('meets.db')
clear()
sheet = load_workbook("source.xlsx").active

i = 0
for merged_cell in sheet.merged_cells:
    try:
        date = extract_date(sheet.cell(row=merged_cell.min_row, column=merged_cell.min_col).value)
    except IndexError:
        continue
    for x in range(merged_cell.min_row, merged_cell.max_row + 1):
        row = list(sheet.rows)[x - 1]
        if row[5].value is None:
            continue
        timeFlag = row[1].value
        if isinstance(timeFlag, datetime.time):
            if timeFlag < datetime.time(hour=12):
                execute(date + timeFlag.strftime("%H:%M"), date + '12:00', row[2].value, row[4].value, row[5].value)
            else:
                execute(date + timeFlag.strftime("%H:%M"), date + '16:30', row[2].value, row[4].value, row[5].value)
        else:
            timeing = timeFlag.split('-')
            if len(timeing[0]) == 4:
                timeing[0] = '0' + timeing[0]
            if len(timeing[1]) == 4:
                timeing[1] = '0' + timeing[1]
            execute(date + timeing[0], date + timeing[1], row[2].value, row[4].value, row[5].value)
